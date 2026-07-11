import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from utils import auto_width


class Exporter:

    @staticmethod
    def _format_worksheet(ws):
        ws.freeze_panes = "A2"

        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = ws.dimensions

        # Bold header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        auto_width(ws)

    @staticmethod
    def export_excel(
        records,
        stats_df,
        summary_df,
        errors_df,
        filename
    ):

        df = pd.DataFrame(records)

        if stats_df is None:
            stats_df = pd.DataFrame()

        if summary_df is None:
            summary_df = pd.DataFrame()

        if errors_df is None:
            errors_df = pd.DataFrame()

        with pd.ExcelWriter(
            filename,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="File Index",
                index=False
            )

            stats_df.to_excel(
                writer,
                sheet_name="Statistics",
                index=False
            )

            summary_df.to_excel(
                writer,
                sheet_name="Summary",
                index=False
            )

            errors_df.to_excel(
                writer,
                sheet_name="Errors",
                index=False
            )

        wb = load_workbook(filename)

        for ws in wb.worksheets:
            Exporter._format_worksheet(ws)

        wb.save(filename)

    @staticmethod
    def export_csv(
        records,
        filename
    ):

        df = pd.DataFrame(records)

        df.to_csv(
            filename,
            index=False,
            encoding="utf-8-sig"
        )